import csv
from datetime import date
from io import BytesIO, StringIO

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.api.routes.employees import require_admin
from app.db.session import get_db
from app.models.user import User
from app.schemas.timesheet import (
    AdminTimesheetItem,
    MyTimesheetItem,
    TimesheetCreateRequest,
    TimesheetCreateResponse,
    TimesheetUpdateRequest,
)
from app.services.timesheets import (
    create_timesheet,
    delete_timesheet,
    get_timesheet,
    list_all_timesheets,
    list_my_timesheets,
    update_timesheet,
)

router = APIRouter()


def _validate_date_range(date_from: date | None, date_to: date | None) -> None:
    if date_from is not None and date_to is not None and date_from > date_to:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="date_from cannot be later than date_to",
        )


def _build_timesheet_workbook(
    rows: list,
    *,
    include_employee: bool,
    title: str,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title

    headers = ["Date"]
    if include_employee:
        headers.extend(["Employee Code", "Employee Name"])
    headers.extend(["Project Code", "Project Name", "Hours", "Overtime", "Remarks"])
    sheet.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="C97A05")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        values = [row.work_date.isoformat()]
        if include_employee:
            values.extend(
                [
                    row.employee.employee_code if row.employee else "",
                    row.employee.full_name if row.employee else "",
                ]
            )
        values.extend(
            [
                row.project.project_code,
                row.project.project_name,
                float(row.hours),
                "Yes" if row.overtime else "No",
                row.remarks or "",
            ]
        )
        sheet.append(values)

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@router.post("", response_model=TimesheetCreateResponse, status_code=status.HTTP_201_CREATED)
def post_timesheet(
    payload: TimesheetCreateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> TimesheetCreateResponse:
    try:
        entry = create_timesheet(db, current_user, payload)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))

    return TimesheetCreateResponse(
        id=entry.id,
        work_date=entry.work_date,
        project_id=entry.project_id,
        hours=float(entry.hours),
        overtime=entry.overtime,
        remarks=entry.remarks,
    )


def _can_manage_timesheet(current_user: User, entry_user_id: int) -> bool:
    return current_user.id == entry_user_id or (
        current_user.role is not None and current_user.role.name == "Admin"
    )


@router.put("/{timesheet_id}", response_model=TimesheetCreateResponse)
def put_timesheet(
    timesheet_id: int,
    payload: TimesheetUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> TimesheetCreateResponse:
    entry = get_timesheet(db, timesheet_id)
    if entry is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Timesheet not found")
    if not _can_manage_timesheet(current_user, entry.user_id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied")
    try:
        entry = update_timesheet(db, entry, payload)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    return TimesheetCreateResponse(
        id=entry.id,
        work_date=entry.work_date,
        project_id=entry.project_id,
        hours=float(entry.hours),
        overtime=entry.overtime,
        remarks=entry.remarks,
    )


@router.delete("/{timesheet_id}", status_code=status.HTTP_204_NO_CONTENT)
def remove_timesheet(
    timesheet_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> None:
    entry = get_timesheet(db, timesheet_id)
    if entry is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Timesheet not found")
    if not _can_manage_timesheet(current_user, entry.user_id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied")
    delete_timesheet(db, entry)


@router.get("/me", response_model=list[MyTimesheetItem])
def get_my_timesheets(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[MyTimesheetItem]:
    rows = list_my_timesheets(db, current_user.id)
    return [
        MyTimesheetItem(
            id=row.id,
            work_date=row.work_date,
            project_id=row.project_id,
            project_code=row.project.project_code,
            project_name=row.project.project_name,
            hours=float(row.hours),
            overtime=row.overtime,
            remarks=row.remarks,
        )
        for row in rows
    ]


@router.get("/me/export.xlsx")
def export_my_timesheets_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Response:
    rows = list_my_timesheets(db, current_user.id)
    content = _build_timesheet_workbook(rows, include_employee=False, title="My Timesheets")
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="my_timesheets.xlsx"'},
    )


@router.get("", response_model=list[AdminTimesheetItem])
def get_all_timesheets(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
) -> list[AdminTimesheetItem]:
    _validate_date_range(date_from, date_to)
    rows = list_all_timesheets(db, date_from=date_from, date_to=date_to)
    return [
        AdminTimesheetItem(
            id=row.id,
            work_date=row.work_date,
            employee_code=row.employee.employee_code if row.employee else None,
            employee_name=row.employee.full_name if row.employee else None,
            project_id=row.project_id,
            project_code=row.project.project_code,
            project_name=row.project.project_name,
            hours=float(row.hours),
            overtime=row.overtime,
            remarks=row.remarks,
        )
        for row in rows
    ]


@router.get("/export")
def export_timesheets_csv(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
) -> Response:
    _validate_date_range(date_from, date_to)
    rows = list_all_timesheets(db, date_from=date_from, date_to=date_to)
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "Date",
            "Employee Code",
            "Employee Name",
            "Project Code",
            "Project Name",
            "Hours",
            "Overtime",
            "Remarks",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                row.work_date.isoformat(),
                row.employee.employee_code if row.employee else "",
                row.employee.full_name if row.employee else "",
                row.project.project_code,
                row.project.project_name,
                float(row.hours),
                "Yes" if row.overtime else "No",
                row.remarks or "",
            ]
        )

    filename = "timesheets.csv"
    if date_from or date_to:
        filename = (
            f"timesheets_{date_from.isoformat() if date_from else 'start'}"
            f"_{date_to.isoformat() if date_to else 'end'}.csv"
        )

    return Response(
        content=buffer.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export.xlsx")
def export_timesheets_excel(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
) -> Response:
    _validate_date_range(date_from, date_to)
    rows = list_all_timesheets(db, date_from=date_from, date_to=date_to)
    content = _build_timesheet_workbook(rows, include_employee=True, title="All Timesheets")
    filename = "timesheets.xlsx"
    if date_from or date_to:
        filename = (
            f"timesheets_{date_from.isoformat() if date_from else 'start'}"
            f"_{date_to.isoformat() if date_to else 'end'}.xlsx"
        )
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
